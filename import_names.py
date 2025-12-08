"""
导入脚本：从 data_get 中的 Excel 文件读取学号/姓名，更新 ticket.db 中 valid_ids 表，
为 valid_ids 增加 student_name 字段（如果不存在）。

用法:
    python import_names.py         # 只更新已存在的 student_id
    python import_names.py --insert-missing   # 对缺失的 student_id 也插入新记录

依赖: openpyxl
"""
import os
import sqlite3
import argparse
from openpyxl import load_workbook

# 默认在仓库根目录运行
DB_PATH = os.path.join(os.path.dirname(__file__), 'ticket.db')
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data_get')

EXCEL_FILES = [
    os.path.join(DATA_DIR, '25.xlsx'),
    os.path.join(DATA_DIR, '242.xlsx'),
    os.path.join(DATA_DIR, '241.xlsx'),
]

FALLBACK_SID_KEYS = ['学号', 'student_id', 'sid', '学号 ']
FALLBACK_NAME_KEYS = ['姓名', 'name', 'student_name', '姓名 '] 


def ensure_column(conn):
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(valid_ids)")
    cols = [row[1] for row in cur.fetchall()]
    if 'student_name' not in cols:
        cur.execute('ALTER TABLE valid_ids ADD COLUMN student_name TEXT')
        conn.commit()
        print('已添加列: student_name')
    else:
        print('列 student_name 已存在')


def find_column_indices(header):
    # header: list of header strings
    sid_idx = None
    name_idx = None
    for i, h in enumerate(header):
        if not h:
            continue
        hh = str(h).strip()
        if sid_idx is None and hh in FALLBACK_SID_KEYS:
            sid_idx = i
        if name_idx is None and hh in FALLBACK_NAME_KEYS:
            name_idx = i
    return sid_idx, name_idx


def read_excel_files(files):
    mapping = {}
    for f in files:
        if not os.path.exists(f):
            print(f'跳过不存在的文件: {f}')
            continue
        print(f'读取文件: {f}')
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = None
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
        else:
            ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            print(f'文件 {f} 没有数据，跳过')
            continue
        sid_idx, name_idx = find_column_indices(header)
        if sid_idx is None or name_idx is None:
            print(f'在文件 {f} 未找到学号或姓名列（尝试的候选项: {FALLBACK_SID_KEYS} / {FALLBACK_NAME_KEYS}），跳过')
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = row[sid_idx] if sid_idx < len(row) else None
            name = row[name_idx] if name_idx < len(row) else None
            if sid is None:
                continue
            sid = str(sid).strip()
            if not sid:
                continue
            if name is None:
                name = ''
            name = str(name).strip()
            # later files overwrite earlier ones
            mapping[sid] = name
    return mapping


def update_db(mapping, insert_missing=False):
    if not mapping:
        print('没有可导入的数据')
        return
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        ensure_column(conn)
        cur = conn.cursor()
        updated = 0
        inserted = 0
        skipped = 0
        for sid, name in mapping.items():
            cur.execute('SELECT 1 FROM valid_ids WHERE student_id = ?', (sid,))
            if cur.fetchone():
                cur.execute('UPDATE valid_ids SET student_name = ? WHERE student_id = ?', (name, sid))
                updated += 1
            else:
                if insert_missing:
                    cur.execute('INSERT INTO valid_ids (student_id, student_name) VALUES (?, ?)', (sid, name))
                    inserted += 1
                else:
                    skipped += 1
        conn.commit()
        print(f'更新完成: updated={updated}, inserted={inserted}, skipped={skipped}')
    finally:
        conn.close()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='从 Excel 导入姓名到 valid_ids 表')
    parser.add_argument('--insert-missing', action='store_true', help='对缺失的 student_id 也插入新记录')
    args = parser.parse_args()

    mapping = read_excel_files(EXCEL_FILES)
    print(f'共解析到 {len(mapping)} 条学号-姓名映射')
    update_db(mapping, insert_missing=args.insert_missing)
